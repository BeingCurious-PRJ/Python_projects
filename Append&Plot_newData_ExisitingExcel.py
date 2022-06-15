import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (LineChart, Reference)
from openpyxl.chart.axis import DateAxis


def append_newData_toExcel(arg1, arg2, arg3, arg4, writer_):
    delta = abs(arg1 - arg2)
    # adding new dataframe
    df_new = pd.Dataframe({'column name1': [arg4],
                           'Column name2': [arg1],
                           'Column name3': [arg2],
                           'Column name4': [delta],
                           'Column name5': [arg3]})
    # appending to existing excel sheet
    df_new.to_excel(writer_, sheet_name='Sheet1', startrow=writer_.sheets['Sheet1'].max_row,
                    index=False, header=False)


def plot_data(ws, first_row_, last_row_):
    # plotting charts in excel with openpyxl
    # chart1
    c1 = LineChart()
    c1.title = "chart title"
    c1.style = 13
    c1.y_axis.title = "Delta[km]"
    c1.y_axis.crossAx = 500
    c1.x_axis = DateAxis(crossAx=100)
    c1.x_axis.number_format = 'd-mmm'
    c1.x_axis.majorTimeUnit = "days"
    c1.x_axis.title = "Date"
    data = Reference(ws, min_col=4, min_row=first_row_ - 1, max_col=4, max_row=last_row_ + 1)
    c1.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=first_row_ - 1, max_row=last_row_ + 1)
    c1.set_categories(dates)
    # styling the lines
    # style1
    s1 = c1.series[0]
    s1.marker.symbol = "diamond"
    s1.marker.graphicalProperties.solidFill = "FF0000"  # marker filling
    s1.marker.graphicalProperties.line.solidFill = "FF0000"  # marker outline
    s1.graphicalProperties.line.noFill = True
    # s2 = c1.series[0]
    # s2.smooth =True
    # chart2 multiple data plot
    c2 = LineChart()
    c2.title = "chart title"
    c2.style = 13
    c2.y_axis_title = "y-title"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'd-mmm'
    c2.x_axis.majorTimeUnit = "days"
    c2.x_axis.title = "Date"
    data = Reference(ws, min_col=2, min_row=first_row_ - 1, max_col=5, max_row=last_row_ + 1)
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=first_row_ - 1, max_row=last_row_ + 1)
    c2.set_categories(dates)
    # styling the lines
    # first set of data MCI CIT
    s2 = c2.series[0]
    s2.marker.symbol = "circle"
    s2.marker.graphicalProperties.solidFill = "B489AC"  # marker filling
    s2.marker.graphicalProperties.line.solidFill = "B489AC"  # marker outline
    s2.graphicalProperties.line.noFill = False
    # second set of data MCI PROD
    s3 = c2.series[1]
    s3.marker.symbol = "circle"
    s3.marker.graphicalProperties.solidFill = "00FF00"  # marker filling
    s3.marker.graphicalProperties.line.solidFill = "00FF00"  # marker outline
    s3.graphicalProperties.line.noFill = False
    # third set of data Delta---keep it small since first plot already has it
    s4 = c2.series[2]
    s4.marker.symbol = "triangle"
    s4.marker.graphicalProperties.solidFill = "DF0000"  # marker filling
    s4.marker.graphicalProperties.line.solidFill = "DF0000"  # marker outline
    s4.graphicalProperties.line.noFill = True
    # fourth set of data HCI
    s5 = c2.series[2]
    s5.marker.symbol = "square"
    s5.marker.graphicalProperties.solidFill = "000000"  # marker filling
    s5.marker.graphicalProperties.line.solidFill = "000000"  # marker outline
    s5.graphicalProperties.line.noFill = False

    ws.add_chart(c1, "F2")
    ws.add_chart(c2, "F19")
    writer.close()


# getting the workbook and sheet ready for appending and plotting
book = load_workbook(r'local path where excel is saved')
writer = pd.ExcelWriter(r'local path where excel is saved', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
startrow = writer.sheets['SheetName'].max_row
sheets = book.sheetnames  # ['Plots']
ws = book[sheets[0]]
last_row = startrow  # 16--or any last number
first_row = startrow - 13  # 3

# function calls to append new data from artifacts to excel and plot them
# give new data inputs arg1, arg2, ... through a function or raw
append_newData_toExcel('arg1', 'arg2', 'arg3', 'arg4', writer)  # function call
plot_data(ws, first_row, last_row)  # function call
